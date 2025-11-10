"""
Excel export functionality for the Torchlight Infinite Price Tracker.
Handles exporting drop statistics to Excel format.
"""

import logging
import time
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..constants import (
    EXCEL_COLUMN_PADDING,
    EXCEL_HEADER_COLOR,
    EXCEL_MAX_COLUMN_WIDTH,
    ITEM_TYPES,
    calculate_fe_per_hour,
    calculate_price_with_tax,
    format_duration,
    get_price_freshness_status,
)
from ..file_manager import FileManager
from ..config_manager import ConfigManager

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles exporting drop data to Excel format."""

    def __init__(self, file_manager: FileManager, config_manager: ConfigManager):
        """
        Initialize the Excel exporter.

        Args:
            file_manager: FileManager instance for data access
            config_manager: ConfigManager instance for settings
        """
        self.file_manager = file_manager
        self.config_manager = config_manager

    def prepare_export_data(self, stats: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Prepare drop data for Excel export.

        Args:
            stats: Statistics dictionary containing drops

        Returns:
            List of drop data dictionaries sorted by category and value
        """
        drops = stats['drops']
        full_table = self.file_manager.load_full_table()
        drop_data = []

        for item_id, count in drops.items():
            if item_id not in full_table:
                continue

            item_data = full_table[item_id]
            item_name = item_data.get("name", item_id)
            item_type = item_data.get("type", "Unknown")
            base_price = item_data.get("price", 0)

            # Apply tax if enabled using centralized function
            item_price = calculate_price_with_tax(
                base_price, item_id, self.config_manager.is_tax_enabled()
            )
            total_value = round(count * item_price, 2)

            # Determine freshness status using helper
            now = time.time()
            last_update = item_data.get("last_update", 0)
            status = get_price_freshness_status(last_update, now)

            drop_data.append({
                'category': item_type,
                'name': item_name,
                'count': count,
                'unit_price': round(item_price, 2),
                'total_value': total_value,
                'status': status
            })

        # Sort by category, then by total value descending
        drop_data.sort(
            key=lambda x: (
                ITEM_TYPES.index(x['category']) if x['category'] in ITEM_TYPES else 999,
                -x['total_value']
            )
        )
        return drop_data

    def write_excel_metadata(self, ws, export_type: str, stats: Dict[str, Any]) -> None:
        """
        Write metadata rows to Excel worksheet.

        Args:
            ws: Worksheet to write to
            export_type: Type of export (e.g., "All Drops" or "Current Map Drops")
            stats: Statistics dictionary
        """
        ws.append([f"Torchlight Infinite Drops Export - {export_type}"])
        ws.append([f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])

        # Format and add time elapsed using helper
        duration = stats['duration']
        time_str = format_duration(duration)
        ws.append([f"Time Elapsed: {time_str}"])

        # Add map count for total stats
        if 'map_count' in stats:
            ws.append([f"Map Count: {stats['map_count']}"])

        # Calculate and add FE/hour using helper
        fe_per_hour = calculate_fe_per_hour(stats['income'], duration)
        ws.append([f"FE/Hour: {round(fe_per_hour, 2)}"])
        ws.append([f"Total Income: {round(stats['income'], 2)} FE"])
        ws.append([])  # Empty row

    def write_excel_data(
        self, ws, drop_data: List[Dict[str, Any]], header_font, header_fill
    ) -> None:
        """
        Write drop data to Excel worksheet.

        Args:
            ws: Worksheet to write to
            drop_data: List of drop data dictionaries
            header_font: Font for header row
            header_fill: Fill for header row
        """
        # Write headers
        headers = ["Category", "Item Name", "Quantity", "Unit Price", "Total Value", "Price Status"]
        ws.append(headers)

        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        current_category = None
        for item in drop_data:
            # Add category separator
            if item['category'] != current_category:
                current_category = item['category']
                ws.append([])  # Empty row before new category

            ws.append([
                item['category'],
                item['name'],
                item['count'],
                item['unit_price'],
                item['total_value'],
                item['status']
            ])

    def auto_adjust_column_widths(self, ws) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet to adjust
        """
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            adjusted_width = min(max_length + EXCEL_COLUMN_PADDING, EXCEL_MAX_COLUMN_WIDTH)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_file(
        self, file_path: str, stats: Dict[str, Any], export_type: str
    ) -> Dict[str, Any]:
        """
        Export drop data to an Excel file.

        Args:
            file_path: Path where the Excel file should be saved
            stats: Statistics dictionary containing drop data
            export_type: Type of export ("All Drops" or "Current Map Drops")

        Returns:
            Dictionary with export summary information

        Raises:
            Exception: If export fails
        """
        # Prepare data
        drop_data = self.prepare_export_data(stats)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Drops Export")
        ws.title = "Drops Export"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color=EXCEL_HEADER_COLOR,
            end_color=EXCEL_HEADER_COLOR,
            fill_type="solid"
        )

        # Write content using helper methods
        self.write_excel_metadata(ws, export_type, stats)
        self.write_excel_data(ws, drop_data, header_font, header_fill)
        self.auto_adjust_column_widths(ws)

        # Save the workbook
        wb.save(file_path)
        logger.info(f"Exported drops to: {file_path}")

        # Build summary information
        duration = stats['duration']
        time_str = format_duration(duration)
        fe_per_hour = calculate_fe_per_hour(stats['income'], duration)

        return {
            'file_path': file_path,
            'total_items': len(drop_data),
            'time_str': time_str,
            'map_count': stats.get('map_count'),
            'total_fe': round(stats['income'], 2),
            'fe_per_hour': round(fe_per_hour, 2)
        }
