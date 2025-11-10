"""
Torchlight Infinite Profit Tracker - PyQt5 Version

This is a refactored version using PyQt5 with better code organization, error handling,
type hints, logging, and thread safety.
"""

import logging
import time
import threading
from typing import Optional, Dict, Any
import sys
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QListWidget, QComboBox, QMessageBox, QDialog, QFrame,
    QFileDialog, QSystemTrayIcon, QMenu, QAction
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QIcon, QCloseEvent

from src.constants import (
    APP_TITLE,
    EXCLUDED_ITEM_ID,
    FILTER_ASHES,
    FILTER_COMPASS,
    FILTER_CURRENCY,
    FILTER_GLOW,
    FILTER_OTHERS,
    ITEM_TYPES,
    LOG_POLL_INTERVAL,
    LOG_FILE_REOPEN_INTERVAL,
    STATUS_FRESH,
    STATUS_OLD,
    STATUS_STALE,
    TIME_FRESH_THRESHOLD,
    TIME_STALE_THRESHOLD,
    UI_FONT_FAMILY,
    UI_FONT_SIZE_LARGE,
    UI_FONT_SIZE_MEDIUM,
    UI_LISTBOX_HEIGHT,
    UI_LISTBOX_WIDTH,
    calculate_price_with_tax,
)
from src.config_manager import ConfigManager
from src.file_manager import FileManager
from src.log_parser import LogParser
from src.inventory_tracker import InventoryTracker
from src.statistics_tracker import StatisticsTracker
from src.game_detector import GameDetector

# Setup logging with UTF-8 encoding to handle Unicode characters
# Note: When running as a GUI application (console=False), sys.stdout may be None
handlers = [logging.FileHandler('tracker.log', encoding='utf-8')]

# Only add console handler if stdout exists (i.e., when running with console)
if sys.stdout is not None:
    stream_handler = logging.StreamHandler(sys.stdout)
    handlers.append(stream_handler)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=handlers
)

# Set the StreamHandler to use UTF-8 encoding (only if stream exists and supports reconfigure)
for handler in logging.root.handlers:
    if isinstance(handler, logging.StreamHandler) and handler.stream is not None:
        try:
            if hasattr(handler.stream, 'reconfigure'):
                handler.stream.reconfigure(encoding='utf-8')
        except (AttributeError, OSError):
            # Ignore if reconfigure fails or isn't supported
            pass

logger = logging.getLogger(__name__)


class WorkerSignals(QObject):
    """Signals for worker thread communication."""
    initialization_complete = pyqtSignal(int)
    update_display = pyqtSignal()
    reshow_drops = pyqtSignal()


class TrackerApp(QMainWindow):
    """Main application window for the Torchlight Infinite Price Tracker."""

    def __init__(
        self,
        config_manager: ConfigManager,
        file_manager: FileManager,
        inventory_tracker: InventoryTracker,
        statistics_tracker: StatisticsTracker,
        log_file_path: Optional[str]
    ):
        """
        Initialize the tracker application.

        Args:
            config_manager: Configuration manager instance.
            file_manager: File manager instance.
            inventory_tracker: Inventory tracker instance.
            statistics_tracker: Statistics tracker instance.
            log_file_path: Path to the game log file.
        """
        super().__init__()

        self.config_manager = config_manager
        self.file_manager = file_manager
        self.inventory_tracker = inventory_tracker
        self.statistics_tracker = statistics_tracker
        self.log_file_path = log_file_path

        self.app_running = True
        self.show_all = False
        self.current_show_types = ITEM_TYPES.copy()

        # Worker signals for thread-safe UI updates
        self.signals = WorkerSignals()
        self.signals.initialization_complete.connect(self.on_initialization_complete)
        self.signals.update_display.connect(self.update_display)
        self.signals.reshow_drops.connect(self.reshow)

        # Initialize color palette
        self.colors = {
            'bg_primary': '#1e1e2e',
            'bg_secondary': '#2a2a3e',
            'bg_tertiary': '#363650',
            'accent': '#8b5cf6',
            'accent_hover': '#9d70f7',
            'success': '#10b981',
            'warning': '#f59e0b',
            'error': '#ef4444',
            'text_primary': '#e2e8f0',
            'text_secondary': '#94a3b8',
            'border': '#4a4a5e',
        }

        self._setup_window()
        self._apply_stylesheet()
        self._create_widgets()

        # Create dialog windows
        self._create_drops_window()
        self._create_settings_window()

        # Create system tray icon
        self._create_tray_icon()

        # Load saved window geometry
        self._load_window_geometry()

    def _setup_window(self) -> None:
        """Setup the main window properties."""
        self.setWindowTitle(APP_TITLE)
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.Window)

        # Make window resizable with minimum size constraints
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.setMinimumSize(400, 600)
        self.resize(500, 800)  # Default size

    def _apply_stylesheet(self) -> None:
        """Apply Qt Style Sheet for modern dark theme."""
        stylesheet = f"""
            QMainWindow {{
                background-color: {self.colors['bg_primary']};
            }}

            QWidget {{
                background-color: {self.colors['bg_primary']};
                color: {self.colors['text_primary']};
                font-family: 'Segoe UI';
                font-size: 11pt;
            }}

            QFrame.card {{
                background-color: {self.colors['bg_tertiary']};
                border-radius: 8px;
                padding: 15px;
            }}

            QLabel {{
                color: {self.colors['text_primary']};
                background-color: transparent;
            }}

            QLabel.header {{
                font-size: 13pt;
                font-weight: bold;
                color: {self.colors['text_primary']};
                padding: 5px;
            }}

            QLabel.status {{
                font-size: 9pt;
                color: {self.colors['text_secondary']};
            }}

            QPushButton {{
                background-color: {self.colors['accent']};
                color: {self.colors['text_primary']};
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
                font-size: 10pt;
            }}

            QPushButton:hover {{
                background-color: {self.colors['accent_hover']};
            }}

            QPushButton:pressed {{
                background-color: {self.colors['accent']};
            }}

            QPushButton.secondary {{
                background-color: {self.colors['bg_secondary']};
                color: {self.colors['text_primary']};
                padding: 6px 12px;
                font-size: 9pt;
                font-weight: normal;
            }}

            QPushButton.secondary:hover {{
                background-color: {self.colors['bg_tertiary']};
            }}

            QPushButton.filter-active {{
                background-color: {self.colors['accent']};
                color: {self.colors['text_primary']};
                padding: 6px 12px;
                font-size: 9pt;
                font-weight: bold;
            }}

            QPushButton.filter-active:hover {{
                background-color: {self.colors['accent_hover']};
            }}

            QPushButton.danger {{
                background-color: {self.colors['error']};
                color: {self.colors['text_primary']};
            }}

            QPushButton.danger:hover {{
                background-color: #dc2626;
            }}

            QPushButton:disabled {{
                background-color: {self.colors['bg_secondary']};
                color: {self.colors['text_secondary']};
            }}

            QListWidget {{
                background-color: {self.colors['bg_secondary']};
                color: {self.colors['text_primary']};
                border: none;
                border-radius: 6px;
                padding: 5px;
                font-size: 10pt;
            }}

            QListWidget::item {{
                padding: 5px;
            }}

            QListWidget::item:selected {{
                background-color: {self.colors['accent']};
                color: {self.colors['text_primary']};
            }}

            QComboBox {{
                background-color: {self.colors['bg_secondary']};
                color: {self.colors['text_primary']};
                border: 1px solid {self.colors['border']};
                border-radius: 4px;
                padding: 5px 10px;
                min-width: 150px;
            }}

            QComboBox:hover {{
                border-color: {self.colors['accent']};
            }}

            QComboBox::drop-down {{
                border: none;
                width: 20px;
            }}

            QComboBox QAbstractItemView {{
                background-color: {self.colors['bg_secondary']};
                color: {self.colors['text_primary']};
                selection-background-color: {self.colors['accent']};
                border: 1px solid {self.colors['border']};
            }}

            QDialog {{
                background-color: {self.colors['bg_primary']};
            }}
        """
        self.setStyleSheet(stylesheet)

    def _create_widgets(self) -> None:
        """Create all GUI widgets."""
        central_widget = self.centralWidget()
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # Stats card
        stats_card = self._create_stats_card()
        main_layout.addWidget(stats_card)

        # Control panel
        control_card = self._create_control_card()
        main_layout.addWidget(control_card)

        # Drops card
        drops_card = self._create_drops_card()
        main_layout.addWidget(drops_card)

    def _create_stats_card(self) -> QFrame:
        """Create the statistics display card."""
        card = QFrame()
        card.setProperty("class", "card")
        card.setObjectName("card")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Map count at the top
        map_count_label = QLabel("MAP COUNT")
        map_count_label.setProperty("class", "status")
        layout.addWidget(map_count_label)

        self.label_map_count = QLabel("🎫 0 maps")
        self.label_map_count.setProperty("class", "header")
        layout.addWidget(self.label_map_count)

        # Separator
        separator1 = QFrame()
        separator1.setFrameShape(QFrame.HLine)
        separator1.setStyleSheet(f"background-color: {self.colors['border']}; max-height: 1px;")
        layout.addWidget(separator1)

        # Current map stats
        current_label = QLabel("CURRENT MAP")
        current_label.setProperty("class", "status")
        layout.addWidget(current_label)

        current_grid = QGridLayout()
        current_grid.setSpacing(15)

        self.label_current_time = QLabel("⏱ 0m00s")
        self.label_current_time.setProperty("class", "header")
        current_grid.addWidget(self.label_current_time, 0, 0)

        self.label_current_speed = QLabel("🔥 0 /min")
        self.label_current_speed.setProperty("class", "header")
        current_grid.addWidget(self.label_current_speed, 0, 1)

        self.label_current_map_fe = QLabel("🔥 0 FE")
        self.label_current_map_fe.setProperty("class", "header")
        current_grid.addWidget(self.label_current_map_fe, 0, 2)

        layout.addLayout(current_grid)

        # Separator
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.HLine)
        separator2.setStyleSheet(f"background-color: {self.colors['border']}; max-height: 1px;")
        layout.addWidget(separator2)

        # Total stats
        total_label = QLabel("TOTAL SESSION")
        total_label.setProperty("class", "status")
        layout.addWidget(total_label)

        total_grid = QGridLayout()
        total_grid.setSpacing(15)

        self.label_total_time = QLabel("⏱ 0m00s")
        self.label_total_time.setProperty("class", "header")
        total_grid.addWidget(self.label_total_time, 0, 0)

        self.label_total_speed = QLabel("🔥 0 /min")
        self.label_total_speed.setProperty("class", "header")
        total_grid.addWidget(self.label_total_speed, 0, 1)

        self.label_total_fe = QLabel("🔥 0 FE")
        self.label_total_fe.setProperty("class", "header")
        total_grid.addWidget(self.label_total_fe, 0, 2)

        layout.addLayout(total_grid)

        return card

    def _create_control_card(self) -> QFrame:
        """Create the control panel card."""
        card = QFrame()
        card.setProperty("class", "card")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Initialization section
        init_label = QLabel("INITIALIZATION")
        init_label.setProperty("class", "status")
        layout.addWidget(init_label)

        init_row = QHBoxLayout()
        init_row.setSpacing(10)

        self.button_initialize = QPushButton("Initialize Tracker")
        self.button_initialize.setCursor(Qt.PointingHandCursor)
        self.button_initialize.clicked.connect(self.start_initialization)
        init_row.addWidget(self.button_initialize)

        self.label_initialize_status = QLabel("Not initialized")
        self.label_initialize_status.setProperty("class", "status")
        init_row.addWidget(self.label_initialize_status)
        init_row.addStretch()

        layout.addLayout(init_row)

        # Separator
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet(f"background-color: {self.colors['border']}; max-height: 1px;")
        layout.addWidget(separator)

        # Actions section
        actions_label = QLabel("ACTIONS")
        actions_label.setProperty("class", "status")
        layout.addWidget(actions_label)

        button_row = QHBoxLayout()
        button_row.setSpacing(5)

        button_log = QPushButton("🔍 Debug Log")
        button_log.setProperty("class", "secondary")
        button_log.setCursor(Qt.PointingHandCursor)
        button_log.clicked.connect(self.debug_log_format)
        button_row.addWidget(button_log)

        button_export = QPushButton("📊 Export")
        button_export.setProperty("class", "secondary")
        button_export.setCursor(Qt.PointingHandCursor)
        button_export.clicked.connect(self.export_drops_to_excel)
        button_row.addWidget(button_export)

        button_settings = QPushButton("⚙ Settings")
        button_settings.setProperty("class", "secondary")
        button_settings.setCursor(Qt.PointingHandCursor)
        button_settings.clicked.connect(self.show_settings_window)
        button_row.addWidget(button_settings)

        layout.addLayout(button_row)

        return card

    def _create_drops_card(self) -> QFrame:
        """Create the drops display card."""
        card = QFrame()
        card.setProperty("class", "card")

        layout = QVBoxLayout(card)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Header
        header_layout = QHBoxLayout()

        drops_title = QLabel("RECENT DROPS")
        drops_title.setProperty("class", "status")
        header_layout.addWidget(drops_title)

        header_layout.addStretch()

        self.button_change = QPushButton("Current Map")
        self.button_change.setProperty("class", "secondary")
        self.button_change.setCursor(Qt.PointingHandCursor)
        self.button_change.clicked.connect(self.change_states)
        header_layout.addWidget(self.button_change)

        layout.addLayout(header_layout)

        # Filter buttons
        filters_label = QLabel("FILTERS")
        filters_label.setProperty("class", "status")
        layout.addWidget(filters_label)

        # First row of filter buttons
        filter_row1 = QHBoxLayout()
        filter_row1.setSpacing(5)

        self.btn_filter_all = QPushButton("All")
        self.btn_filter_all.setProperty("class", "filter-active")
        self.btn_filter_all.setCursor(Qt.PointingHandCursor)
        self.btn_filter_all.clicked.connect(lambda: self.set_filter(ITEM_TYPES))
        filter_row1.addWidget(self.btn_filter_all)

        self.btn_filter_currency = QPushButton("Currency")
        self.btn_filter_currency.setProperty("class", "secondary")
        self.btn_filter_currency.setCursor(Qt.PointingHandCursor)
        self.btn_filter_currency.clicked.connect(lambda: self.set_filter(FILTER_CURRENCY))
        filter_row1.addWidget(self.btn_filter_currency)

        self.btn_filter_embers = QPushButton("Embers")
        self.btn_filter_embers.setProperty("class", "secondary")
        self.btn_filter_embers.setCursor(Qt.PointingHandCursor)
        self.btn_filter_embers.clicked.connect(lambda: self.set_filter(FILTER_ASHES))
        filter_row1.addWidget(self.btn_filter_embers)

        layout.addLayout(filter_row1)

        # Second row of filter buttons
        filter_row2 = QHBoxLayout()
        filter_row2.setSpacing(5)

        self.btn_filter_compass = QPushButton("Compass")
        self.btn_filter_compass.setProperty("class", "secondary")
        self.btn_filter_compass.setCursor(Qt.PointingHandCursor)
        self.btn_filter_compass.clicked.connect(lambda: self.set_filter(FILTER_COMPASS))
        filter_row2.addWidget(self.btn_filter_compass)

        self.btn_filter_memory = QPushButton("Memory")
        self.btn_filter_memory.setProperty("class", "secondary")
        self.btn_filter_memory.setCursor(Qt.PointingHandCursor)
        self.btn_filter_memory.clicked.connect(lambda: self.set_filter(FILTER_GLOW))
        filter_row2.addWidget(self.btn_filter_memory)

        self.btn_filter_others = QPushButton("Others")
        self.btn_filter_others.setProperty("class", "secondary")
        self.btn_filter_others.setCursor(Qt.PointingHandCursor)
        self.btn_filter_others.clicked.connect(lambda: self.set_filter(FILTER_OTHERS))
        filter_row2.addWidget(self.btn_filter_others)

        layout.addLayout(filter_row2)

        # Store filter buttons for easy access
        self.filter_buttons = {
            tuple(ITEM_TYPES): self.btn_filter_all,
            tuple(FILTER_CURRENCY): self.btn_filter_currency,
            tuple(FILTER_ASHES): self.btn_filter_embers,
            tuple(FILTER_COMPASS): self.btn_filter_compass,
            tuple(FILTER_GLOW): self.btn_filter_memory,
            tuple(FILTER_OTHERS): self.btn_filter_others,
        }

        # Drops list
        self.inner_panel_drop_listbox = QListWidget()
        self.inner_panel_drop_listbox.setMinimumHeight(UI_LISTBOX_HEIGHT * 20)  # Approximate height
        self.inner_panel_drop_listbox.addItem("Drops will be displayed here...")
        layout.addWidget(self.inner_panel_drop_listbox)

        return card

    def _create_drops_window(self) -> None:
        """Create the drops detail dialog."""
        self.drops_dialog = QDialog(self)
        self.drops_dialog.setWindowTitle("Drops Detail - FurTorch")
        self.drops_dialog.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.Tool)
        self.drops_dialog.setModal(False)

        layout = QVBoxLayout(self.drops_dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # Header card
        header_card = QFrame()
        header_card.setProperty("class", "card")
        header_layout = QVBoxLayout(header_card)
        header_layout.setContentsMargins(15, 15, 15, 15)

        header_label = QLabel("DROP FILTERS")
        header_label.setProperty("class", "status")
        header_layout.addWidget(header_label)

        self.button_toggle_all = QPushButton("Current: Current Map Drops (Click to toggle All Drops)")
        self.button_toggle_all.setCursor(Qt.PointingHandCursor)
        self.button_toggle_all.clicked.connect(self.change_states)
        header_layout.addWidget(self.button_toggle_all)

        layout.addWidget(header_card)

        # Filter buttons card
        filters_card = QFrame()
        filters_card.setProperty("class", "card")
        filters_layout = QVBoxLayout(filters_card)
        filters_layout.setContentsMargins(15, 15, 15, 15)

        filters_label = QLabel("ITEM CATEGORIES")
        filters_label.setProperty("class", "status")
        filters_layout.addWidget(filters_label)

        filter_items = [
            ("All Items", ITEM_TYPES),
            ("Currency", FILTER_CURRENCY),
            ("Embers", FILTER_ASHES),
            ("Compass", FILTER_COMPASS),
            ("Memory", FILTER_GLOW),
            ("Others", FILTER_OTHERS)
        ]

        for text, filter_type in filter_items:
            btn = QPushButton(text)
            btn.setProperty("class", "secondary")
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda checked, f=filter_type: self.set_filter(f))
            filters_layout.addWidget(btn)

        layout.addWidget(filters_card)

    def _create_settings_window(self) -> None:
        """Create the settings dialog."""
        self.settings_dialog = QDialog(self)
        self.settings_dialog.setWindowTitle("Settings - FurTorch")
        self.settings_dialog.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.Tool)
        self.settings_dialog.setModal(False)

        layout = QVBoxLayout(self.settings_dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # Tax settings card
        tax_card = QFrame()
        tax_card.setProperty("class", "card")
        tax_layout = QVBoxLayout(tax_card)
        tax_layout.setContentsMargins(15, 15, 15, 15)

        tax_title = QLabel("TAX SETTINGS")
        tax_title.setProperty("class", "status")
        tax_layout.addWidget(tax_title)

        tax_row = QHBoxLayout()
        label_tax = QLabel("Price Calculation:")
        tax_row.addWidget(label_tax)

        config = self.config_manager.get()
        self.tax_combo = QComboBox()
        self.tax_combo.addItems(["No tax", "Include tax"])
        self.tax_combo.setCurrentIndex(config.tax)
        self.tax_combo.currentIndexChanged.connect(self.change_tax)
        tax_row.addWidget(self.tax_combo)

        tax_layout.addLayout(tax_row)
        layout.addWidget(tax_card)

        # Actions card
        actions_card = QFrame()
        actions_card.setProperty("class", "card")
        actions_layout = QVBoxLayout(actions_card)
        actions_layout.setContentsMargins(15, 15, 15, 15)

        actions_title = QLabel("ACTIONS")
        actions_title.setProperty("class", "status")
        actions_layout.addWidget(actions_title)

        reset_button = QPushButton("Reset Statistics")
        reset_button.setProperty("class", "danger")
        reset_button.setCursor(Qt.PointingHandCursor)
        reset_button.clicked.connect(self.reset_tracking)
        actions_layout.addWidget(reset_button)

        layout.addWidget(actions_card)

    def start_initialization(self) -> None:
        """Start the inventory initialization process."""
        if self.inventory_tracker.start_initialization():
            self.label_initialize_status.setText("Waiting for bag update...")
            self.label_initialize_status.setStyleSheet(f"color: {self.colors['warning']};")
            self.button_initialize.setEnabled(False)

            QMessageBox.information(
                self,
                "Initialization",
                "Click 'OK' and then sort your bag in-game by clicking the sort button.\n\n"
                "This will refresh your inventory and allow the tracker to initialize "
                "with the correct item counts."
            )
        else:
            QMessageBox.information(
                self,
                "Initialization",
                "Initialization already in progress. Please wait."
            )

    def on_initialization_complete(self, item_count: int) -> None:
        """
        Called when initialization completes.

        Args:
            item_count: Number of unique items initialized.
        """
        self.label_initialize_status.setText(f"✓ Initialized ({item_count} items)")
        self.label_initialize_status.setStyleSheet(f"color: {self.colors['success']};")
        self.button_initialize.setEnabled(True)

    def closeEvent(self, event: QCloseEvent) -> None:
        """Handle window close event."""
        reply = QMessageBox.question(
            self,
            "Exit",
            "Are you sure you want to exit?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Signal the app to stop
            self.app_running = False

            # Hide tray icon
            if hasattr(self, 'tray_icon') and self.tray_icon:
                self.tray_icon.hide()

            # Close child dialogs
            try:
                self.drops_dialog.close()
            except (AttributeError, RuntimeError) as e:
                logger.debug(f"Error closing drops dialog: {e}")

            try:
                self.settings_dialog.close()
            except (AttributeError, RuntimeError) as e:
                logger.debug(f"Error closing settings dialog: {e}")

            # Accept the event and quit the application
            event.accept()
            QApplication.quit()
        else:
            event.ignore()

    def reset_tracking(self) -> None:
        """Reset tracking statistics while preserving inventory initialization."""
        reply = QMessageBox.question(
            self,
            "Reset Statistics",
            "Are you sure you want to reset all tracking statistics? "
            "This will clear all drop statistics and map counts.\n\n"
            "Your inventory initialization will be preserved.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.statistics_tracker.reset()

            # Update UI
            self.label_total_fe.setText("🔥 0 FE")
            self.label_current_map_fe.setText("🔥 0 FE")
            self.label_map_count.setText("🎫 0 maps")
            self.inner_panel_drop_listbox.clear()

            QMessageBox.information(
                self,
                "Reset Complete",
                "Statistics have been reset. Inventory initialization preserved."
            )

    def change_tax(self, value: int) -> None:
        """
        Change the tax setting.

        Args:
            value: 0 for no tax, 1 for tax enabled.
        """
        self.config_manager.update_tax(value)
        logger.info(f"Tax setting changed to: {'Enabled' if value else 'Disabled'}")

    def change_states(self) -> None:
        """Toggle between current map and all drops view."""
        self.show_all = not self.show_all
        if not self.show_all:
            self.button_toggle_all.setText("Current: Current Map Drops (Click to toggle All Drops)")
            self.button_change.setText("Current Map")
        else:
            self.button_toggle_all.setText("Current: All Drops (Click to toggle Current Map Drops)")
            self.button_change.setText("All Drops")
        self.reshow()

    def set_filter(self, item_types: list) -> None:
        """
        Set the item type filter.

        Args:
            item_types: List of item types to display.
        """
        self.current_show_types = item_types

        # Update button styling to show active filter
        filter_key = tuple(item_types)
        for key, button in self.filter_buttons.items():
            if key == filter_key:
                button.setProperty("class", "filter-active")
            else:
                button.setProperty("class", "secondary")
            # Force style update
            button.style().unpolish(button)
            button.style().polish(button)

        self.reshow()

    def show_drops_window(self) -> None:
        """Toggle the drops detail window."""
        if self.drops_dialog.isVisible():
            self.drops_dialog.hide()
        else:
            self.drops_dialog.show()

    def show_settings_window(self) -> None:
        """Toggle the settings window."""
        if self.settings_dialog.isVisible():
            self.settings_dialog.hide()
        else:
            self.settings_dialog.show()

    def debug_log_format(self) -> None:
        """Print debug information about current state."""
        logger.info("=== DEBUG INFO ===")
        logger.info(f"Bag initialized: {self.inventory_tracker.bag_initialized}")
        logger.info(f"Initialization complete: {self.inventory_tracker.initialization_complete}")

        bag_summary = self.inventory_tracker.get_bag_state_summary()
        logger.info(f"Total items tracked: {len(bag_summary)}")

        full_table = self.file_manager.load_full_table()
        for item_id, total in bag_summary.items():
            name = full_table.get(item_id, {}).get("name", f"Unknown (ID: {item_id})")
            logger.info(f"  {name}: {total}")

        QMessageBox.information(
            self,
            "Debug Information",
            f"Debug information has been logged.\n\n"
            f"Bag initialized: {self.inventory_tracker.bag_initialized}\n"
            f"Initialization complete: {self.inventory_tracker.initialization_complete}\n"
            f"Total items tracked: {len(bag_summary)}"
        )

    def reshow(self) -> None:
        """Refresh the drop display."""
        full_table = self.file_manager.load_full_table()

        # Get appropriate drop list
        if self.show_all:
            stats = self.statistics_tracker.get_total_stats()
        else:
            stats = self.statistics_tracker.get_current_map_stats()

        # Update labels
        current_map_stats = self.statistics_tracker.get_current_map_stats()
        self.label_current_map_fe.setText(f"🔥 {round(current_map_stats['income'], 2)} FE")

        total_stats = self.statistics_tracker.get_total_stats()
        self.label_total_fe.setText(f"🔥 {round(total_stats['income'], 2)} FE")
        self.label_map_count.setText(f"🎫 {total_stats['map_count']} maps")

        # Prepare drop items with their values for sorting
        drop_items = []
        for item_id, count in stats['drops'].items():
            if item_id not in full_table:
                continue

            item_data = full_table[item_id]
            item_name = item_data.get("name", item_id)
            item_type = item_data.get("type", "Unknown")

            if item_type not in self.current_show_types:
                continue

            # Determine status based on last update time
            now = time.time()
            last_update = item_data.get("last_update", 0)
            time_passed = now - last_update

            if time_passed < TIME_FRESH_THRESHOLD:
                status = STATUS_FRESH
            elif time_passed < TIME_STALE_THRESHOLD:
                status = STATUS_STALE
            else:
                status = STATUS_OLD

            # Calculate price with tax if applicable using centralized function
            base_price = item_data.get("price", 0)
            item_price = calculate_price_with_tax(base_price, item_id, self.config_manager.is_tax_enabled())

            total_value = round(count * item_price, 2)
            drop_items.append((total_value, f"{status} {item_name} x{count} [{total_value}]"))

        # Sort by total value (descending - highest first)
        drop_items.sort(key=lambda x: x[0], reverse=True)

        # Update drop listbox
        self.inner_panel_drop_listbox.clear()
        for _, item_text in drop_items:
            self.inner_panel_drop_listbox.addItem(item_text)

    def update_display(self) -> None:
        """Update the time and income displays."""
        if self.statistics_tracker.is_in_map:
            current_stats = self.statistics_tracker.get_current_map_stats()
            duration = current_stats['duration']

            m = int(duration // 60)
            s = int(duration % 60)
            self.label_current_time.setText(f"⏱ {m}m{s:02d}s")

            income_per_min = current_stats['income_per_minute']
            self.label_current_speed.setText(f"🔥 {round(income_per_min, 2)} /min")

            # Update current map FE
            self.label_current_map_fe.setText(f"🔥 {round(current_stats['income'], 2)} FE")

        total_stats = self.statistics_tracker.get_total_stats()
        duration = total_stats['duration']
        m = int(duration // 60)
        s = int(duration % 60)
        self.label_total_time.setText(f"⏱ {m}m{s:02d}s")

        income_per_min = total_stats['income_per_minute']
        self.label_total_speed.setText(f"🔥 {round(income_per_min, 2)} /min")

        # Update total session FE
        self.label_total_fe.setText(f"🔥 {round(total_stats['income'], 2)} FE")

        # Update map count
        self.label_map_count.setText(f"🎫 {total_stats['map_count']} maps")

    def _create_tray_icon(self) -> None:
        """Create the system tray icon."""
        # Create tray icon (using default application icon for now)
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(self.style().standardIcon(self.style().SP_ComputerIcon))

        # Create tray menu
        tray_menu = QMenu()

        show_action = QAction("Show", self)
        show_action.triggered.connect(self.show_from_tray)
        tray_menu.addAction(show_action)

        hide_action = QAction("Hide", self)
        hide_action.triggered.connect(self.hide)
        tray_menu.addAction(hide_action)

        tray_menu.addSeparator()

        quit_action = QAction("Quit", self)
        quit_action.triggered.connect(self.quit_application)
        tray_menu.addAction(quit_action)

        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.on_tray_icon_activated)
        self.tray_icon.show()

        logger.info("System tray icon created")

    def show_from_tray(self) -> None:
        """Show window from tray."""
        self.show()
        self.activateWindow()
        self.raise_()

    def on_tray_icon_activated(self, reason: QSystemTrayIcon.ActivationReason) -> None:
        """Handle tray icon activation."""
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_from_tray()

    def quit_application(self) -> None:
        """Quit the application."""
        self.app_running = False
        QApplication.quit()

    def changeEvent(self, event) -> None:
        """Handle window state changes."""
        if event.type() == event.WindowStateChange:
            if self.isMinimized():
                # Minimize to tray
                QTimer.singleShot(0, self.hide)
                self.tray_icon.showMessage(
                    "Torchlight Price Tracker",
                    "Application minimized to tray",
                    QSystemTrayIcon.Information,
                    2000
                )
        super().changeEvent(event)

    def moveEvent(self, event) -> None:
        """Handle window move event."""
        super().moveEvent(event)
        # Save window geometry after move
        self._save_window_geometry()

    def resizeEvent(self, event) -> None:
        """Handle window resize event."""
        super().resizeEvent(event)
        # Save window geometry after resize
        self._save_window_geometry()

    def _load_window_geometry(self) -> None:
        """Load saved window geometry from config."""
        config = self.config_manager.get()
        if config.window_x is not None and config.window_y is not None:
            if config.window_width is not None and config.window_height is not None:
                self.setGeometry(config.window_x, config.window_y,
                                config.window_width, config.window_height)
                logger.info(f"Loaded window geometry: {config.window_x},{config.window_y} "
                           f"{config.window_width}x{config.window_height}")

    def _save_window_geometry(self) -> None:
        """Save window geometry to config."""
        if not self.isMaximized() and not self.isMinimized():
            geometry = self.geometry()
            self.config_manager.update_window_geometry(
                geometry.x(), geometry.y(),
                geometry.width(), geometry.height()
            )

    def export_drops_to_excel(self) -> None:
        """Export drops to an Excel file sorted by item category."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self,
                "Missing Dependency",
                "The openpyxl library is not installed.\n\n"
                "Please install it using: pip install openpyxl"
            )
            return

        try:
            # Get the appropriate drop list based on current view
            if self.show_all:
                stats = self.statistics_tracker.get_total_stats()
                export_type = "All Drops"
            else:
                stats = self.statistics_tracker.get_current_map_stats()
                export_type = "Current Map Drops"

            drops = stats['drops']

            if not drops:
                QMessageBox.information(
                    self,
                    "No Drops",
                    "There are no drops to export."
                )
                return

            full_table = self.file_manager.load_full_table()

            # Prepare data sorted by category
            drop_data = []
            for item_id, count in drops.items():
                if item_id not in full_table:
                    continue

                item_data = full_table[item_id]
                item_name = item_data.get("name", item_id)
                item_type = item_data.get("type", "Unknown")
                base_price = item_data.get("price", 0)

                # Apply tax if enabled using centralized function
                item_price = calculate_price_with_tax(base_price, item_id, self.config_manager.is_tax_enabled())

                total_value = round(count * item_price, 2)

                # Determine freshness status
                now = time.time()
                last_update = item_data.get("last_update", 0)
                time_passed = now - last_update

                if time_passed < TIME_FRESH_THRESHOLD:
                    status = "Fresh"
                elif time_passed < TIME_STALE_THRESHOLD:
                    status = "Stale"
                else:
                    status = "Old"

                drop_data.append({
                    'category': item_type,
                    'name': item_name,
                    'count': count,
                    'unit_price': round(item_price, 2),
                    'total_value': total_value,
                    'status': status
                })

            # Sort by category, then by total value descending
            drop_data.sort(key=lambda x: (ITEM_TYPES.index(x['category']) if x['category'] in ITEM_TYPES else 999, -x['total_value']))

            # Prompt user for save location
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"torchlight_drops_{timestamp}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                default_filename,
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return  # User cancelled

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Drops Export"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
            category_font = Font(bold=True, size=11)
            category_fill = PatternFill(start_color="2a2a3e", end_color="2a2a3e", fill_type="solid")

            # Write metadata
            ws.append([f"Torchlight Infinite Drops Export - {export_type}"])
            ws.append([f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])

            # Calculate time elapsed
            duration = stats['duration']
            hours = int(duration // 3600)
            minutes = int((duration % 3600) // 60)
            seconds = int(duration % 60)
            time_str = f"{hours}h {minutes}m {seconds}s" if hours > 0 else f"{minutes}m {seconds}s"
            ws.append([f"Time Elapsed: {time_str}"])

            # Add map count for total stats
            if 'map_count' in stats:
                ws.append([f"Map Count: {stats['map_count']}"])

            # Calculate and add FE/hour
            fe_per_hour = (stats['income'] / (duration / 3600)) if duration > 0 else 0
            ws.append([f"FE/Hour: {round(fe_per_hour, 2)}"])

            ws.append([f"Total Income: {round(stats['income'], 2)} FE"])
            ws.append([])  # Empty row

            # Write headers
            headers = ["Category", "Item Name", "Quantity", "Unit Price", "Total Value", "Price Status"]
            ws.append(headers)

            header_row = ws[ws.max_row]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data
            current_category = None
            for item in drop_data:
                # Add category separator
                if item['category'] != current_category:
                    current_category = item['category']
                    ws.append([])  # Empty row before new category

                ws.append([
                    item['category'],
                    item['name'],
                    item['count'],
                    item['unit_price'],
                    item['total_value'],
                    item['status']
                ])

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(file_path)

            # Build success message with all stats
            success_msg = f"Drops have been exported to:\n{file_path}\n\n"
            success_msg += f"Total items: {len(drop_data)}\n"
            success_msg += f"Time elapsed: {time_str}\n"
            if 'map_count' in stats:
                success_msg += f"Map count: {stats['map_count']}\n"
            success_msg += f"Total FE: {round(stats['income'], 2)}\n"
            success_msg += f"FE/Hour: {round(fe_per_hour, 2)}"

            QMessageBox.information(
                self,
                "Export Successful",
                success_msg
            )
            logger.info(f"Drops exported to: {file_path}")

        except Exception as e:
            logger.error(f"Error exporting drops to Excel: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Export Error",
                f"An error occurred while exporting:\n{str(e)}"
            )


class LogMonitorThread(threading.Thread):
    """Thread for monitoring the game log file."""

    def __init__(
        self,
        app: TrackerApp,
        log_file_path: Optional[str],
        log_parser: LogParser,
        inventory_tracker: InventoryTracker,
        statistics_tracker: StatisticsTracker,
        signals: WorkerSignals
    ):
        """
        Initialize the log monitor thread.

        Args:
            app: Main application instance.
            log_file_path: Path to the game log file.
            log_parser: Log parser instance.
            inventory_tracker: Inventory tracker instance.
            statistics_tracker: Statistics tracker instance.
            signals: Worker signals for thread-safe UI updates.
        """
        super().__init__(daemon=True)
        self.app = app
        self.log_file_path = log_file_path
        self.log_parser = log_parser
        self.inventory_tracker = inventory_tracker
        self.statistics_tracker = statistics_tracker
        self.signals = signals
        self.log_file = None
        self.last_reopen_check = time.time()

    def _open_log_file(self) -> bool:
        """
        Open the log file and seek to end.

        Returns:
            True if successful, False otherwise.
        """
        if not self.log_file_path:
            return False

        try:
            self.log_file = open(self.log_file_path, "r", encoding="utf-8")
            self.log_file.seek(0, 2)  # Seek to end
            logger.info("Log file opened successfully")
            return True
        except (IOError, OSError) as e:
            logger.error(f"Could not open log file: {e}")
            self.log_file = None
            return False

    def _close_log_file(self) -> None:
        """Close the log file if open."""
        if self.log_file:
            try:
                self.log_file.close()
                logger.info("Log file closed")
            except (IOError, OSError) as e:
                logger.error(f"Error closing log file: {e}")
            finally:
                self.log_file = None

    def _check_and_reopen_log_file(self) -> None:
        """Check if log file needs reopening (e.g., was deleted or rotated)."""
        now = time.time()
        if now - self.last_reopen_check < LOG_FILE_REOPEN_INTERVAL:
            return

        self.last_reopen_check = now

        # Check if file still exists and is accessible
        if self.log_file_path:
            import os
            if not os.path.exists(self.log_file_path):
                logger.warning("Log file no longer exists, attempting to reopen")
                self._close_log_file()
                self._open_log_file()

    def run(self) -> None:
        """Run the log monitoring loop."""
        if not self.log_file_path:
            logger.warning("No log file path provided")
            return

        # Open log file initially
        self._open_log_file()

        try:
            while self.app.app_running:
                try:
                    # Sleep in smaller chunks to be more responsive to shutdown
                    for _ in range(int(LOG_POLL_INTERVAL * 10)):
                        if not self.app.app_running:
                            break
                        time.sleep(0.1)

                    if not self.app.app_running:
                        break

                    # Check if log file needs reopening
                    self._check_and_reopen_log_file()

                    # Read and process log file
                    if self.log_file:
                        try:
                            text = self.log_file.read()
                            if text:
                                self._process_log_text(text)
                        except (IOError, OSError) as e:
                            logger.error(f"Error reading log file: {e}")
                            # Try to reopen the file
                            self._close_log_file()
                            self._open_log_file()

                    # Update display via signal
                    self.signals.update_display.emit()

                except Exception as e:
                    logger.error(f"Error in log monitor thread: {e}", exc_info=True)

        finally:
            # Ensure log file is closed on exit
            self._close_log_file()

    def _process_log_text(self, text: str) -> None:
        """
        Process new log text.

        Args:
            text: New log text to process.
        """
        # Update prices
        self.log_parser.update_prices_in_table(text)

        # Check for initialization completion
        if self.inventory_tracker.awaiting_initialization:
            success, item_count = self.inventory_tracker.process_initialization(text)
            if success:
                self.signals.initialization_complete.emit(item_count)

        # Detect map changes
        entering_map, exiting_map = self.log_parser.detect_map_change(text)

        if entering_map:
            self.statistics_tracker.enter_map()
            self.inventory_tracker.reset_map_baseline()

        if exiting_map:
            self.statistics_tracker.exit_map()

        # Detect item changes
        changes = self.inventory_tracker.scan_for_changes(text)
        if changes:
            self.statistics_tracker.process_item_changes(changes)
            self.signals.reshow_drops.emit()

            # If not in map but got changes, assume we're in a map
            if not self.statistics_tracker.is_in_map:
                self.statistics_tracker.is_in_map = True


def main():
    """Main entry point for the application."""
    logger.info("=== Torchlight Infinite Price Tracker Starting (PyQt5) ===")

    # Create QApplication
    app = QApplication(sys.argv)

    # Initialize managers
    config_manager = ConfigManager()
    file_manager = FileManager()

    # Initialize data files
    file_manager.ensure_file_exists("config.json", {"opacity": 1.0, "tax": 0, "user": ""})
    file_manager.ensure_file_exists("translation_mapping.json", {})
    file_manager.initialize_full_table_from_en_table()

    # Initialize core components
    log_parser = LogParser(file_manager)
    inventory_tracker = InventoryTracker(log_parser)
    statistics_tracker = StatisticsTracker(file_manager, config_manager)

    # Detect game and log file
    game_detector = GameDetector()
    game_found, log_file_path = game_detector.detect_game()

    # Create and run the application
    tracker_app = TrackerApp(
        config_manager,
        file_manager,
        inventory_tracker,
        statistics_tracker,
        log_file_path
    )

    # Show game not found warning after app is created
    if not game_found:
        logger.warning("Game not found - tracker will run without log monitoring")
        QTimer.singleShot(100, lambda: QMessageBox.warning(
            tracker_app,
            "Game Not Found",
            "Could not find Torchlight: Infinite game process or log file.\n\n"
            "The tool will continue running but won't be able to track drops "
            "until the game is started.\n\n"
            "Please make sure the game is running with logging enabled, "
            "then restart this tool."
        ))

    # Start log monitoring thread
    monitor = LogMonitorThread(
        tracker_app,
        log_file_path,
        log_parser,
        inventory_tracker,
        statistics_tracker,
        tracker_app.signals
    )
    monitor.start()

    # Show the window
    tracker_app.show()

    # Run the application
    logger.info("Application started")
    exit_code = app.exec_()
    logger.info("Application shut down")
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
